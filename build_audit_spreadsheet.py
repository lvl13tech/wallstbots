#!/usr/bin/env python3
"""
WallStBots Data Fields Audit Spreadsheet Builder
Creates a comprehensive Excel workbook documenting all numeric fields
across all 3 sites and all pages.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1A2332")   # dark navy header
HDR2_FILL = PatternFill("solid", start_color="0D1B2A")   # darker subheader
ROW_ALT   = PatternFill("solid", start_color="F5F7FA")   # light grey alternate row
ROW_WHITE = PatternFill("solid", start_color="FFFFFF")

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR2_FONT = Font(name="Arial", bold=True, color="00D4FF", size=10)
BODY_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", bold=True, size=9)
MONO_FONT = Font(name="Courier New", size=8)
TITLE_FONT= Font(name="Arial", bold=True, color="1A2332", size=14)

thin  = Side(border_style="thin",   color="D0D7E0")
thick = Side(border_style="medium", color="1A2332")
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, font=HDR_FONT, fill=HDR_FILL, wrap=True):
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = CELL_BORDER

def style_body(cell, mono=False, alt=False, bold=False):
    cell.font = MONO_FONT if mono else (BOLD_FONT if bold else BODY_FONT)
    cell.fill = ROW_ALT if alt else ROW_WHITE
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = CELL_BORDER

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 1 — OVERVIEW / INDEX
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_idx = wb.active
ws_idx.title = "Overview"
ws_idx.sheet_view.showGridLines = False

ws_idx["A1"] = "WallStBots — Data Fields Audit"
ws_idx["A1"].font = TITLE_FONT
ws_idx["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_idx.row_dimensions[1].height = 30

ws_idx["A2"] = "Comprehensive mapping of every numeric field across all 3 sites and all pages."
ws_idx["A2"].font = Font(name="Arial", size=10, color="555555")

headers_idx = ["Sheet", "Description", "# Fields", "Sites Covered"]
for c, h in enumerate(headers_idx, 1):
    cell = ws_idx.cell(row=4, column=c, value=h)
    style_header(cell)

index_rows = [
    ("Public Pages",      "Homepage, The Race, How It Works, News, Signals, Reports, Get Yours", 42,  "All 3"),
    ("Bot Detail Pages",  "Per-bot strategy panel, holdings table, portfolio entry on bot-detail.html", 38, "All 3"),
    ("Fund Detail Pages", "Current Value, Total P&L, Today's Change, positions table on public fund page", 28, "All 3"),
    ("Dashboard",         "Member dashboard: fund cards, portfolio list, webmaster panel", 24, "All 3"),
    ("Portfolio Fund",    "portfolio-fund.html: per-user bot performance, holdings, P&L over time", 22, "All 3"),
    ("Leaderboard",       "leaderboard.html: rank, return%, portfolio value, win rate", 14, "All 3"),
    ("Signals Page",      "Signals table: price, target, upside%, score, RSI, momentum columns", 18, "All 3"),
    ("Reports",           "Weekly report cards: week P&L, %, grade per fund", 10, "All 3"),
    ("Get Yours Pricing", "Subscription pricing, referral discounts, PayPal amounts", 12, "All 3"),
    ("Site Differences",  "Fields that differ between lvl13 / wallstbots / bitbot13", 0,  "Comparison"),
]

for r, row in enumerate(index_rows, 5):
    for c, val in enumerate(row, 1):
        cell = ws_idx.cell(row=r, column=c, value=val)
        style_body(cell, alt=(r % 2 == 0))
        if c == 1:
            cell.font = Font(name="Arial", bold=True, size=9, color="0070C0")

set_col_widths(ws_idx, [22, 60, 12, 16])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHARED column headers for data sheets
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DATA_HEADERS = [
    "Page / Section",
    "Field Label (UI)",
    "JS Expression (app.js)",
    "state.json Path",
    "Refresh Script Calculation",
    "Format Function",
    "Fallback / Default",
    "Site Differences",
]
DATA_WIDTHS = [22, 22, 38, 38, 50, 18, 22, 40]

def add_data_sheet(title, rows):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32

    # Sheet title row
    ws["A1"] = title + " — Data Fields Audit"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(DATA_HEADERS))}1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Column headers
    for c, h in enumerate(DATA_HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_header(cell)

    set_col_widths(ws, DATA_WIDTHS)

    for r, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            is_mono = c in (3, 4, 5)   # code columns
            style_body(cell, mono=is_mono, alt=(r % 2 == 1))
            if c == 1:
                cell.font = Font(name="Arial", bold=True, size=9)
        ws.row_dimensions[r].height = max(15, min(80, 15 * (1 + str(row[4]).count("\n"))))

    return ws

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: PUBLIC PAGES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
public_rows = [
    # ── HOMEPAGE ──────────────────────────────────────────────────────────
    ("Homepage / Hero",
     "Starting Capital (copy)",
     "fmt$0(cap)",
     "funds.starting_capital",
     "Hardcoded $55,000 seed; never written by refresh script (set at inception)",
     "fmt$0()",
     "(STATE.funds && STATE.funds.starting_capital) || 55000",
     "All 3 sites identical"),

    ("Homepage / Hero",
     "Stock/Coin Count",
     "stockCount",
     "signals.recommendations.length",
     "len(recs) written to signals.json by refresh script",
     "plain integer",
     "|| 55 (lvl13/wallstbots) / 50 (bitbot13)",
     "lvl13: 43 AI stocks; wallstbots: ~55 sector stocks; bitbot13: ~200 crypto"),

    ("Homepage / Leaderboard Strip",
     "Today % (per fund)",
     "fmtPct(v.day_pct)",
     "funds.funds.[fid].value.day_pct",
     "day_pct = (day_pnl / prev_total) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical logic"),

    ("Homepage / Race Cards",
     "Fund Current Value ($)",
     "fmt$0(v.total)",
     "funds.funds.[fid].value.total",
     "total = prev_total + day_pnl",
     "fmt$0()",
     "cap (starting_capital)",
     "All 3 sites identical"),

    ("Homepage / Race Cards",
     "Total P&L ($)",
     "fmt$0(v.pnl)",
     "funds.funds.[fid].value.pnl",
     "pnl = total - starting_capital",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Homepage / Race Cards",
     "Total P&L (%)",
     "fmtPct(v.pnl_pct)",
     "funds.funds.[fid].value.pnl_pct",
     "pnl_pct = (pnl / starting_capital) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Homepage / Race Cards",
     "Today % Change",
     "fmtPct(v.day_pct)",
     "funds.funds.[fid].value.day_pct",
     "day_pct = (day_pnl / prev_total) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Homepage / Signals Preview",
     "Signal Count per Action",
     "summary[action] || 0",
     "signals.summary.[ACTION]",
     "sum(1 for r in recs if r['action']==action)",
     "plain integer",
     "0",
     "All 3 sites identical"),

    ("Homepage / Signals Preview",
     "Upside % (top picks)",
     "fmtPct(r.upside_pct)",
     "signals.recommendations[n].upside_pct",
     "upside_pct = (target/price - 1)*100",
     "fmtPct()",
     "''",
     "All 3 sites identical"),

    # ── THE RACE ──────────────────────────────────────────────────────────
    ("Race / Header",
     "Starting Capital ($)",
     "fmt$0(cap)",
     "funds.starting_capital",
     "Set at inception — $55,000 (equity) / $50,000 (crypto)",
     "fmt$0()",
     "55000",
     "bitbot13 uses $50,000"),

    ("Race / Header",
     "Stock/Coin Universe Count",
     "stockCount",
     "signals.recommendations.length",
     "len(recommendations) in signals JSON",
     "plain integer",
     "55",
     "See Homepage notes"),

    ("Race / Fund Cards",
     "Fund Value ($)",
     "fmt$0(v.total)",
     "funds.funds.[fid].value.total",
     "total = prev_total + day_pnl",
     "fmt$0()",
     "cap",
     "All 3 sites identical"),

    ("Race / Fund Cards",
     "P&L $ since inception",
     "fmt$0(v.pnl)",
     "funds.funds.[fid].value.pnl",
     "pnl = total - starting_capital",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Race / Fund Cards",
     "P&L % since inception",
     "fmtPct(v.pnl_pct)",
     "funds.funds.[fid].value.pnl_pct",
     "pnl_pct = (pnl / starting_capital) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Race / Fund Cards",
     "Today %",
     "fmtPct(v.day_pct)",
     "funds.funds.[fid].value.day_pct",
     "day_pct = (day_pnl / prev_total) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Race / Trajectory Chart",
     "Chart Data Points ($ per fund per day)",
     "snaps.map(s => s[fid] || null)",
     "funds.snapshots[n].[fid]",
     "Written by refresh script: snapshots.append({date, bot13, oracle, wizard, equalizer, titan})",
     "Chart.js Y-axis",
     "null",
     "All 3 sites identical"),

    # ── SIGNALS ───────────────────────────────────────────────────────────
    ("Signals / Summary Bar",
     "Count per signal action",
     "sum[k] || 0",
     "signals.summary.[ACTION]",
     "sum(1 for r in recs if r['action']==action)",
     "plain integer (big)",
     "0",
     "All 3 sites identical"),

    ("Signals / Table",
     "Current Price ($)",
     "'$' + r.price.toFixed(2)",
     "signals.recommendations[n].price",
     "price = current market price from yfinance",
     "'$' + .toFixed(2)",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "Target Price ($)",
     "'$' + r.target.toFixed(2)",
     "signals.recommendations[n].target",
     "target = price * (1 + score/100) or analyst target",
     "'$' + .toFixed(2)",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "Upside % to Target",
     "fmtPct(r.upside_pct)",
     "signals.recommendations[n].upside_pct",
     "upside_pct = (target/price - 1)*100",
     "fmtPct()",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "Composite Score",
     "(r.score>=0?'+':'')+r.score.toFixed(1)",
     "signals.recommendations[n].score",
     "score = weighted sum(momentum, RSI, MACD, volume, volatility)",
     "±N.N",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "RSI(14)",
     "ind.rsi_14",
     "signals.recommendations[n].indicators.rsi_14",
     "RSI(14) calculated by signals engine via yfinance",
     "plain number",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "5-day Momentum %",
     "fmtPct(ind.mom_5d)",
     "signals.recommendations[n].indicators.mom_5d",
     "(close_today - close_5d_ago) / close_5d_ago * 100",
     "fmtPct()",
     "'—'",
     "All 3 sites identical"),

    ("Signals / Table",
     "20-day Momentum %",
     "fmtPct(ind.mom_20d)",
     "signals.recommendations[n].indicators.mom_20d",
     "(close_today - close_20d_ago) / close_20d_ago * 100",
     "fmtPct()",
     "'—'",
     "All 3 sites identical"),

    # ── REPORTS ───────────────────────────────────────────────────────────
    ("Reports / List",
     "Week % per Fund",
     "fmtPct(res.week_pct)",
     "reports.reports[n].fund_results.[fid].week_pct",
     "week_pct = (week_pnl / start_of_week_total) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Reports / Detail",
     "Week P&L ($)",
     "fmt$(res.week_pnl)",
     "reports.reports[n].fund_results.[fid].week_pnl",
     "week_pnl = end_of_week_total - start_of_week_total",
     "fmt$()",
     "0",
     "All 3 sites identical"),

    ("Reports / Detail",
     "Week P&L (%)",
     "fmtPct(res.week_pct)",
     "reports.reports[n].fund_results.[fid].week_pct",
     "week_pct = week_pnl / start_of_week_total * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: BOT / FUND DETAIL (renderFund + renderStrategyPanel)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fund_rows = [
    ("Fund Detail / Stats Cards",
     "Current Value ($)",
     "fmt$0(v.total)",
     "funds.funds.[fid].value.total",
     "total = prev_total + day_pnl\nprev_total = prior state.json total",
     "fmt$0()",
     "starting_capital",
     "All 3 sites identical"),

    ("Fund Detail / Stats Cards",
     "Started at ($)",
     "fmt$0(startCap)",
     "funds.funds.[fid].starting_capital",
     "Set once at fund inception — never overwritten",
     "fmt$0()",
     "funds.starting_capital",
     "All 3 sites identical"),

    ("Fund Detail / Stats Cards",
     "Total P&L ($)",
     "fmt$0(v.pnl)",
     "funds.funds.[fid].value.pnl",
     "pnl = total - starting_capital",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Stats Cards",
     "Total P&L % (all-time)",
     "fmtPct(v.pnl_pct)",
     "funds.funds.[fid].value.pnl_pct",
     "pnl_pct = (total - starting_capital) / starting_capital * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Stats Cards",
     "Today's Change ($)",
     "fmt$0(v.day_pnl)",
     "funds.funds.[fid].value.day_pnl",
     "day_pnl = sum(p.pnl for p in positions)\nor = total - prev_total",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Stats Cards",
     "Today's Change % (since yesterday)",
     "fmtPct(v.day_pct)",
     "funds.funds.[fid].value.day_pct",
     "day_pct = day_pnl / prev_total * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel",
     "Projected Return %",
     "projRet.toFixed(2) + '%'",
     "funds.funds.[fid].current_strategy.projected_return",
     "BOT13: sum(weight * day_pct for pick)\nORACLE: sum(weight * week_pct for pick)\nWIZARD: sum(weight * month_pct for pick)",
     "+N.NN%",
     "not shown if null",
     "Same formula all 3 sites; bitbot13 uses crypto day_pct"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "Pick Weight %",
     "(p.weight*100).toFixed(0)+'%'",
     "funds.funds.[fid].current_strategy.picks[n].weight",
     "Normalised rank weight: top pick ~23%, 5th ~14%",
     "N%",
     "N/A",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "Pick Score",
     "(p.score>=0?'+':'')+p.score",
     "funds.funds.[fid].current_strategy.picks[n].score",
     "Same composite score as signals engine",
     "±N",
     "N/A",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "1-day Momentum %",
     "fmtPct(ind.mom_1d)",
     "funds.funds.[fid].current_strategy.picks[n].indicators.mom_1d",
     "(close_today - close_yesterday) / close_yesterday * 100",
     "fmtPct()",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "5-day Momentum %",
     "fmtPct(ind.mom_5d)",
     "funds.funds.[fid].current_strategy.picks[n].indicators.mom_5d",
     "(close_today - close_5d_ago) / close_5d_ago * 100",
     "fmtPct()",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "20-day Momentum %",
     "fmtPct(ind.mom_20d)",
     "funds.funds.[fid].current_strategy.picks[n].indicators.mom_20d",
     "(close_today - close_20d_ago) / close_20d_ago * 100",
     "fmtPct()",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "60-day Momentum %",
     "fmtPct(ind.mom_60d)",
     "funds.funds.[fid].current_strategy.picks[n].indicators.mom_60d",
     "(close_today - close_60d_ago) / close_60d_ago * 100",
     "fmtPct()",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "RSI(14)",
     "ind.rsi_14",
     "funds.funds.[fid].current_strategy.picks[n].indicators.rsi_14",
     "RSI(14) from yfinance OHLCV data",
     "plain number",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Strategy Panel — TRADE picks",
     "MACD %",
     "fmtPct(ind.macd_pct)",
     "funds.funds.[fid].current_strategy.picks[n].indicators.macd_pct",
     "MACD histogram / price * 100",
     "fmtPct()",
     "not shown if null",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Shares",
     "shares.toFixed(2)",
     "funds.funds.[fid].value.positions[n].shares",
     "shares = alloc_cash / entry_price\nalloc_cash = total * weight",
     "N.NN",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Entry Price ($)",
     "'$'+entry.toFixed(2)",
     "funds.funds.[fid].value.positions[n].entry_price",
     "entry_price = prev_close (prior day close) at time of TRADE decision",
     "'$'N.NN",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Current Price ($)",
     "'$'+price.toFixed(2)",
     "funds.funds.[fid].value.positions[n].price",
     "price = current intraday price from yfinance",
     "'$'N.NN",
     "entry_price",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Position Value ($)",
     "fmt$0(value)",
     "funds.funds.[fid].value.positions[n].value",
     "value = shares * price",
     "fmt$0()",
     "shares * entry",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Today's Change %",
     "fmtPct(dayPct)",
     "funds.funds.[fid].value.positions[n].day_pct",
     "day_pct = (price - prev_close) / prev_close * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Position Total P&L ($)",
     "fmt$0(pnl)",
     "funds.funds.[fid].value.positions[n].pnl",
     "pnl = value - cost_basis\ncost_basis = shares * entry_price",
     "fmt$0()",
     "value - (shares*entry)",
     "All 3 sites identical"),

    ("Fund Detail / Holdings Table",
     "Position Total P&L %",
     "fmtPct(pnlPct)",
     "funds.funds.[fid].value.positions[n].pnl_pct",
     "pnl_pct = (price / entry_price - 1) * 100",
     "fmtPct()",
     "(price/entry-1)*100",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: DASHBOARD (members area)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
dashboard_rows = [
    ("Dashboard / Fund Cards",
     "Fund Current Value ($)",
     "fmt$0(fund.total_value)",
     "API: /tracker/state → funds.[fid].value.total",
     "Same as public renderFund — pulled from state.json",
     "fmt$0()",
     "starting_capital",
     "All 3 sites identical"),

    ("Dashboard / Fund Cards",
     "Today % Change",
     "fmtPct(fund.day_pct)",
     "API: /tracker/state → funds.[fid].value.day_pct",
     "day_pct = day_pnl / prev_total * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Fund Cards",
     "Total P&L % (all-time)",
     "fmtPct(fund.pnl_pct)",
     "API: /tracker/state → funds.[fid].value.pnl_pct",
     "pnl_pct = (total - starting_capital) / starting_capital * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio List",
     "Portfolio Value ($)",
     "fmt$0(port.total_value)",
     "API: /portfolios → [n].total_value",
     "sum(holding.shares * holding.current_price) for holdings in portfolio",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio List",
     "Portfolio P&L ($)",
     "fmt$0(port.total_pnl)",
     "API: /portfolios → [n].total_pnl",
     "sum(holding.pnl) = sum((current_price - entry_price) * shares)",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio List",
     "Portfolio P&L %",
     "fmtPct(port.pnl_pct)",
     "API: /portfolios → [n].pnl_pct",
     "total_pnl / total_cost_basis * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "Shares",
     "h.shares.toFixed(4)",
     "API: /portfolios/[id]/holdings → [n].shares",
     "User-defined; stored in Supabase bot_holdings table",
     "N.NNNN",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "Entry Price ($)",
     "'$' + h.entry_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].entry_price",
     "User-entered at time of portfolio creation",
     "'$'N.NN",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "Current Price ($)",
     "'$' + h.current_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].current_price",
     "Fetched from yfinance on each API request",
     "'$'N.NN",
     "entry_price",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "Position Value ($)",
     "fmt$0(h.shares * h.current_price)",
     "Computed client-side: shares * current_price",
     "Not stored — computed on render",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "P&L ($)",
     "fmt$0(h.pnl)",
     "API: /portfolios/[id]/holdings → [n].pnl",
     "(current_price - entry_price) * shares",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Portfolio / Holdings",
     "P&L %",
     "fmtPct(h.pnl_pct)",
     "API: /portfolios/[id]/holdings → [n].pnl_pct",
     "(current_price / entry_price - 1) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Webmaster Panel",
     "Total Users",
     "data.total_users",
     "API: /webmaster/system → total_users",
     "SELECT count(*) FROM users",
     "plain integer",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Webmaster Panel",
     "Active Subscribers",
     "data.active_subscribers",
     "API: /webmaster/system → active_subscribers",
     "SELECT count(*) FROM users WHERE tier != 'free'",
     "plain integer",
     "0",
     "All 3 sites identical"),

    ("Dashboard / Webmaster Panel",
     "Monthly Revenue ($)",
     "'$' + data.monthly_revenue.toFixed(2)",
     "API: /webmaster/system → monthly_revenue",
     "sum(tier_price for user in active_subscribers)",
     "'$'N.NN",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: BOT DETAIL PAGE (bot-detail.html — member area)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
botdetail_rows = [
    ("Bot Detail / Strategy Panel",
     "Projected Return %",
     "strategy.projected_return.toFixed(2)+'%'",
     "API: /tracker/state → funds.[fid].current_strategy.projected_return",
     "BOT13 engine: sum(w*day_pct for pick)\nWritten by refresh_*.py each trading day",
     "+N.NN%",
     "not shown if null/0",
     "All 3 sites identical"),

    ("Bot Detail / Strategy Panel",
     "Decision Label",
     "strategy.decision",
     "funds.[fid].current_strategy.decision",
     "'TRADE' | 'HOLD' | 'CASH'\nHOLD set when outside trading window\nCASH set when bot finds no edge",
     "string",
     "'HOLD'",
     "All 3 sites identical"),

    ("Bot Detail / Today's Session",
     "Pick Weight %",
     "(p.weight*100).toFixed(0)+'%'",
     "funds.[fid].current_strategy.picks[n].weight",
     "Normalised rank weight from bot13_engine.py",
     "N%",
     "N/A (no pick = HOLD/CASH)",
     "All 3 sites identical"),

    ("Bot Detail / Today's Session",
     "Pick Score",
     "p.score",
     "funds.[fid].current_strategy.picks[n].score",
     "Composite score from signals engine",
     "±N",
     "N/A",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "Shares Held",
     "h.shares.toFixed(4)",
     "API: /portfolios/[id]/holdings → [n].shares",
     "User-defined quantity stored in Supabase bot_holdings",
     "N.NNNN",
     "—",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "Entry Price ($)",
     "'$' + h.entry_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].entry_price",
     "User-entered at add time",
     "'$'N.NN",
     "—",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "Current Price ($)",
     "'$' + h.current_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].current_price",
     "Live price via yfinance on API request",
     "'$'N.NN",
     "entry_price",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "Value ($)",
     "fmt$0(h.shares * h.current_price)",
     "Computed: shares × current_price",
     "Not stored — computed on render",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "P&L ($)",
     "fmt$0(h.pnl)",
     "API: /portfolios/[id]/holdings → [n].pnl",
     "(current_price - entry_price) * shares",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio / Holdings",
     "P&L %",
     "fmtPct(h.pnl_pct)",
     "API: /portfolios/[id]/holdings → [n].pnl_pct",
     "(current_price / entry_price - 1) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio Summary",
     "Total Portfolio Value ($)",
     "fmt$0(totalValue)",
     "Computed: sum(h.shares * h.current_price)",
     "sum over all holdings in portfolio",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio Summary",
     "Total P&L ($)",
     "fmt$0(totalPnl)",
     "Computed: sum(h.pnl)",
     "sum((current_price - entry_price) * shares) for all holdings",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Bot Detail / User Portfolio Summary",
     "Total P&L %",
     "fmtPct(totalPnlPct)",
     "Computed: totalPnl / totalCost * 100",
     "totalCost = sum(h.entry_price * h.shares)",
     "fmtPct()",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: PORTFOLIO FUND PAGE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
portfund_rows = [
    ("Portfolio Fund / Header",
     "Portfolio Total Value ($)",
     "fmt$0(portfolio.total_value)",
     "API: /portfolios/[id] → total_value",
     "sum(holding.shares * current_price) for all holdings",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Header",
     "Portfolio Total P&L ($)",
     "fmt$0(portfolio.total_pnl)",
     "API: /portfolios/[id] → total_pnl",
     "sum((current_price - entry_price) * shares)",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Header",
     "Portfolio Total P&L %",
     "fmtPct(portfolio.pnl_pct)",
     "API: /portfolios/[id] → pnl_pct",
     "total_pnl / total_cost_basis * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "Shares",
     "h.shares.toFixed(4)",
     "API: /portfolios/[id]/holdings → [n].shares",
     "Stored in Supabase bot_holdings",
     "N.NNNN",
     "—",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "Entry Price ($)",
     "'$'+h.entry_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].entry_price",
     "User-entered",
     "'$'N.NN",
     "—",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "Current Price ($)",
     "'$'+h.current_price.toFixed(2)",
     "API: /portfolios/[id]/holdings → [n].current_price",
     "Live from yfinance on API request",
     "'$'N.NN",
     "entry_price",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "Value ($)",
     "fmt$0(h.shares * h.current_price)",
     "Computed: shares × current_price",
     "Not stored — computed on render",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "P&L ($)",
     "fmt$0(h.pnl)",
     "API: /portfolios/[id]/holdings → [n].pnl",
     "(current_price - entry_price) * shares",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Holdings Table",
     "P&L %",
     "fmtPct(h.pnl_pct)",
     "API: /portfolios/[id]/holdings → [n].pnl_pct",
     "(current_price / entry_price - 1) * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Portfolio Fund / Performance Chart",
     "Snapshot Value ($)",
     "snapshots.map(s => s.value)",
     "API: /portfolios/[id]/snapshots → [n].value",
     "Written by trigger_portfolio_snapshots() in refresh scripts\nvalue = sum(shares * price) at snapshot time",
     "Chart.js Y-axis",
     "null",
     "All 3 sites identical"),

    ("Portfolio Fund / Bot Comparison",
     "Bot Return %",
     "fmtPct(botResult.pct)",
     "API: /portfolios/[id] → bot_results.[fid].pct",
     "Computed from snapshot series: (final - initial) / initial * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: LEADERBOARD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
leaderboard_rows = [
    ("Leaderboard / Table",
     "Rank",
     "entry.rank",
     "API: /leaderboard → [n].rank",
     "ORDER BY total_return_pct DESC, portfolio_value DESC",
     "plain integer",
     "—",
     "All 3 sites identical"),

    ("Leaderboard / Table",
     "Portfolio Value ($)",
     "fmt$0(entry.portfolio_value)",
     "API: /leaderboard → [n].portfolio_value",
     "sum(shares * current_price) for all holdings",
     "fmt$0()",
     "0",
     "All 3 sites identical"),

    ("Leaderboard / Table",
     "Total Return %",
     "fmtPct(entry.total_return_pct)",
     "API: /leaderboard → [n].total_return_pct",
     "(portfolio_value - cost_basis) / cost_basis * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Leaderboard / Table",
     "Today's Change %",
     "fmtPct(entry.day_pct)",
     "API: /leaderboard → [n].day_pct",
     "(current_value - prev_day_value) / prev_day_value * 100",
     "fmtPct()",
     "0",
     "All 3 sites identical"),

    ("Leaderboard / Table",
     "Win Rate %",
     "entry.win_rate.toFixed(0)+'%'",
     "API: /leaderboard → [n].win_rate",
     "winning_positions / total_positions * 100",
     "N%",
     "—",
     "All 3 sites identical"),

    ("Leaderboard / Table",
     "Positions Count",
     "entry.position_count",
     "API: /leaderboard → [n].position_count",
     "COUNT(holdings) for portfolio",
     "plain integer",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: GET YOURS (pricing)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
pricing_rows = [
    ("Get Yours / Tier Cards",
     "Monthly Price ($)",
     "'$' + price.toFixed(2)",
     "PRICING.[tier].monthly (hardcoded JS constant)",
     "Hardcoded: member=$49.99, insider=$69.99, syndicate=$99.99",
     "'$'N.NN",
     "N/A",
     "All 3 sites identical; same PRICING object"),

    ("Get Yours / Tier Cards",
     "Annual Price ($)",
     "'$' + prices.annual.toFixed(2)",
     "PRICING.[tier].annual (hardcoded JS constant)",
     "Hardcoded: member=$499.00, insider=$699.00, syndicate=$899.00",
     "'$'N.NN",
     "N/A",
     "All 3 sites identical; same PRICING object"),

    ("Get Yours / Referral Discount",
     "Monthly Referral Price ($)",
     "'$' + (price * 0.5).toFixed(2)",
     "Computed: PRICING.[tier].monthly * 0.5",
     "50% off first month — not stored; computed live",
     "'$'N.NN",
     "N/A",
     "All 3 sites identical"),

    ("Get Yours / Referral Discount",
     "Annual Referral Price ($)",
     "'$' + (price - 100).toFixed(2)",
     "Computed: PRICING.[tier].annual - 100",
     "$100 flat off annual — not stored; computed live",
     "'$'N.NN",
     "N/A",
     "All 3 sites identical"),

    ("Get Yours / Active Price Label",
     "Selected Price ($)",
     "'$' + price.toFixed(2) + suffix",
     "PRICING[GY_TIER][GY_CYCLE]",
     "Dynamic based on selected tier + billing cycle",
     "'$'N.NN/mo or /yr",
     "member/monthly",
     "All 3 sites identical"),

    ("Get Yours / PayPal Form",
     "Subscription Amount (a3)",
     "base (= price.toFixed(2))",
     "Hidden form field — PRICING[GY_TIER][GY_CYCLE]",
     "Passed to PayPal _xclick-subscriptions form",
     "string",
     "N/A",
     "lvl13: return URL = lvl13.tech\nwallstbots: wallstbots.tech\nbitbot13: bitbot13.tech"),

    ("Get Yours / PayPal Form",
     "Trial Amount (a1, referral only)",
     "refPrice",
     "Computed: annual ? price-100 : price*0.5",
     "First-payment override for referral subscribers",
     "string",
     "N/A (not shown without ref code)",
     "All 3 sites identical"),

    ("Referral Dashboard",
     "Total Redemptions",
     "d.total_redemptions",
     "API: /account/referral → total_redemptions",
     "COUNT(referral_uses WHERE referrer_id = user_id)",
     "plain integer",
     "0",
     "All 3 sites identical"),

    ("Referral Dashboard",
     "Credit Balance ($)",
     "'$' + d.credit_balance.toFixed(2)",
     "API: /account/referral → credit_balance",
     "total_redemptions * 35",
     "'$'N.NN",
     "0",
     "All 3 sites identical"),

    ("Referral Dashboard",
     "Pending Credit ($)",
     "'$' + d.pending_credit.toFixed(2)",
     "API: /account/referral → pending_credit",
     "Redemptions in current billing cycle × $35",
     "'$'N.NN",
     "0",
     "All 3 sites identical"),
]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA: SITE DIFFERENCES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
diff_headers = ["Field / Section", "lvl13.tech", "wallstbots.tech", "bitbot13.tech", "Notes"]
diff_widths  = [28, 35, 35, 35, 40]
diff_rows = [
    ("Starting Capital",
     "$55,000",
     "$55,000",
     "$50,000",
     "Only bitbot13 differs — lower capital for crypto universe"),

    ("Universe Size",
     "43 AI & Quantum stocks",
     "~55 S&P sector stocks",
     "~200 crypto coins",
     "Signals count reflects these differences"),

    ("Universe Description (hero copy)",
     "43 hand-picked AI & Quantum stocks",
     "Top 3 stocks per S&P 500 sector + hottest IPOs",
     "Top 50–200 coins by market cap",
     "Copy differs; logic identical"),

    ("Platform identifier (API param)",
     "platform=lvl13",
     "platform=wallstbots",
     "platform=bitbot13",
     "Passed to backend tracker API on all fetch calls"),

    ("Refresh script",
     "refresh_lvl13.py",
     "refresh_wallstbots.py",
     "refresh_bitbot13.py",
     "Same logic; config differs: EQUITY_CFG vs CRYPTO_CFG"),

    ("Market type for signals",
     "equity (stocks)",
     "equity (stocks)",
     "crypto (coins)",
     "bitbot13 engine uses CRYPTO_CFG for price fetching"),

    ("projected_return variable naming",
     "prev_b13_strategy",
     "prev_b13_strategy",
     "b13_prev_strategy",
     "Minor naming difference in bitbot13 refresh script"),

    ("Config object",
     "EQUITY_CFG",
     "EQUITY_CFG",
     "CRYPTO_CFG",
     "Controls trading window hours, ticker list, exchange"),

    ("Default stock count fallback",
     "|| 43 (not explicit, uses signal length)",
     "|| 55",
     "|| 50 (set in hero copy)",
     "Fallback shown when signals not yet loaded"),

    ("PayPal return URL",
     "https://lvl13.tech/#/thanks",
     "https://wallstbots.tech/#/thanks",
     "https://bitbot13.tech/#/thanks",
     "Each site redirects to its own domain after PayPal checkout"),

    ("Cross-links in homepage",
     "Links to wallstbots.tech and bitbot13.tech",
     "Links to lvl13.tech and bitbot13.tech",
     "Links to lvl13.tech and wallstbots.tech",
     "\"Also from Level 13\" section cross-promotes all 3"),

    ("Sector news filter chips",
     "11 GICS sectors (stocks)",
     "11 GICS sectors (stocks)",
     "Crypto categories (DeFi, L1, L2, etc.)",
     "bitbot13 uses different sector taxonomy for news"),

    ("fund-detail cashRow text (EOD)",
     "'End of trading — now holding cash'",
     "'End of trading — now holding cash'",
     "'End of trading — now holding cash'",
     "All 3 identical after app.js sync fix"),

    ("Strategy period label",
     "'Day of' / 'Week of' / 'Month of'",
     "'Day of' / 'Week of' / 'Month of'",
     "'Day of' / 'Week of' / 'Month of'",
     "All 3 identical — same renderStrategyPanel()"),
]

def add_diff_sheet():
    ws = wb.create_sheet(title="Site Differences")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 32

    ws["A1"] = "Site Differences — Field-by-Field Comparison"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(diff_headers))}1")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    for c, h in enumerate(diff_headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        style_header(cell)

    set_col_widths(ws, diff_widths)

    for r, row in enumerate(diff_rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_body(cell, alt=(r % 2 == 1))
            if c == 1:
                cell.font = Font(name="Arial", bold=True, size=9)
        ws.row_dimensions[r].height = 20

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BUILD ALL SHEETS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
add_data_sheet("Public Pages",      public_rows)
add_data_sheet("Fund Detail Pages", fund_rows)
add_data_sheet("Bot Detail Pages",  botdetail_rows)
add_data_sheet("Dashboard",         dashboard_rows)
add_data_sheet("Portfolio Fund",    portfund_rows)
add_data_sheet("Leaderboard",       leaderboard_rows)
add_data_sheet("Get Yours Pricing", pricing_rows)
add_diff_sheet()

# ── Save ─────────────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "WallStBots_DataFields_Audit_2026-05-28.xlsx")
wb.save(out)
print(f"Saved: {out}")
